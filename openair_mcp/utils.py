"""Utility functions for building input contracts from common data sources."""
from __future__ import annotations

import csv
import io
import os
from pathlib import Path
from typing import Sequence

from openair_mcp.contracts import SeriesColumnV1, SeriesMetaV1, SeriesV1, WindSeriesV1
from openair_mcp.datetime_parse import parse_series_timestamp
from openair_mcp.text_encoding import normalize_label

# Max raw file size for load_series_from_upload (override: OPENAIR_INGEST_MAX_BYTES).
INGEST_MAX_BYTES = int(os.getenv("OPENAIR_INGEST_MAX_BYTES", "1048576"))


def _dedupe_parsed_rows(
    parsed_rows: list[tuple[str, dict[str, float | None]]],
) -> list[tuple[str, dict[str, float | None]]]:
    """Keep the last row when duplicate UTC timestamps appear (common in ES network exports)."""
    by_ts: dict[str, dict[str, float | None]] = {}
    order: list[str] = []
    for ts, values in parsed_rows:
        if ts not in by_ts:
            order.append(ts)
        by_ts[ts] = values
    return [(ts, by_ts[ts]) for ts in order]


def _rows_to_series(
    parsed_rows: list[tuple[str, dict[str, float | None]]],
    numeric_cols: list[str],
    timezone: str,
    source: str,
    site: str | None,
    lat: float | None = None,
    lon: float | None = None,
) -> SeriesV1:
    if not parsed_rows:
        raise ValueError("No data rows found")
    timestamps = [ts for ts, _ in parsed_rows]
    col_data: dict[str, list[float | None]] = {col: [] for col in numeric_cols}
    for _, values in parsed_rows:
        for col in numeric_cols:
            col_data[col].append(values.get(col))
    series = [
        SeriesColumnV1(name=normalize_label(col), unit="", values=vals)
        for col, vals in col_data.items()
    ]
    return SeriesV1(
        timestamps=timestamps,
        series=series,
        meta=SeriesMetaV1(source=source, site=site, timezone=timezone, lat=lat, lon=lon),
    )


def _normalize_table_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    normalized: list[dict[str, str]] = []
    for row in rows:
        normalized.append({normalize_label(k): v for k, v in row.items()})
    return normalized


def _parse_table_rows(
    rows: list[dict[str, str]],
    datetime_col: str,
    columns: Sequence[str] | None,
    timezone: str,
    dedupe_timestamps: bool,
) -> tuple[list[str], list[tuple[str, dict[str, float | None]]]]:
    if not rows:
        raise ValueError("Table is empty")
    rows = _normalize_table_rows(rows)
    datetime_col = normalize_label(datetime_col)
    if datetime_col not in rows[0]:
        raise ValueError(f"Column '{datetime_col}' not found. Available: {list(rows[0].keys())}")

    if columns:
        numeric_cols = [normalize_label(c) for c in columns]
        missing = [c for c in numeric_cols if c not in rows[0]]
        if missing:
            raise ValueError(f"Columns not found: {missing}. Available: {list(rows[0].keys())}")
    else:
        numeric_cols = [c for c in rows[0] if c != datetime_col]

    parsed_rows: list[tuple[str, dict[str, float | None]]] = []
    for row in rows:
        ts = parse_series_timestamp(row[datetime_col], source_timezone=timezone)
        values: dict[str, float | None] = {}
        for col in numeric_cols:
            raw = (row.get(col) or "").strip()
            try:
                values[col] = float(raw)
            except (ValueError, TypeError):
                values[col] = None
        parsed_rows.append((ts, values))

    if dedupe_timestamps:
        parsed_rows = _dedupe_parsed_rows(parsed_rows)
    return numeric_cols, parsed_rows


def _excel_raw_rows_to_dict_rows(raw_rows: list[tuple]) -> list[dict[str, str]]:
    if not raw_rows:
        raise ValueError("Excel sheet is empty")
    headers = [
        normalize_label(str(h).strip()) if h is not None else ""
        for h in raw_rows[0]
    ]
    dict_rows: list[dict[str, str]] = []
    for raw in raw_rows[1:]:
        if raw is None or all(cell is None or str(cell).strip() == "" for cell in raw):
            continue
        row: dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            cell = raw[idx] if idx < len(raw) else None
            row[header] = "" if cell is None else str(cell)
        dict_rows.append(row)
    return dict_rows


def series_from_csv_text(
    text: str,
    datetime_col: str = "date",
    columns: Sequence[str] | None = None,
    timezone: str = "UTC",
    source: str = "file",
    site: str | None = None,
    lat: float | None = None,
    lon: float | None = None,
    dedupe_timestamps: bool = True,
) -> SeriesV1:
    """Build SeriesV1 from CSV text (UTF-8 with optional BOM)."""
    rows = list(csv.DictReader(io.StringIO(text)))
    numeric_cols, parsed_rows = _parse_table_rows(
        rows, datetime_col, columns, timezone, dedupe_timestamps
    )
    return _rows_to_series(parsed_rows, numeric_cols, timezone, source, site, lat, lon)


def series_from_excel_bytes(
    data: bytes,
    sheet_name: str | None = None,
    datetime_col: str = "date",
    columns: Sequence[str] | None = None,
    timezone: str = "UTC",
    source: str = "file",
    site: str | None = None,
    lat: float | None = None,
    lon: float | None = None,
    dedupe_timestamps: bool = True,
) -> SeriesV1:
    """Build SeriesV1 from an in-memory .xlsx workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("openpyxl is required for Excel import. pip install openpyxl") from exc

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    dict_rows = _excel_raw_rows_to_dict_rows(raw_rows)
    numeric_cols, parsed_rows = _parse_table_rows(
        dict_rows, datetime_col, columns, timezone, dedupe_timestamps
    )
    return _rows_to_series(parsed_rows, numeric_cols, timezone, source, site, lat, lon)


def series_from_csv(
    path: str | Path,
    datetime_col: str = "date",
    columns: Sequence[str] | None = None,
    timezone: str = "UTC",
    source: str = "file",
    site: str | None = None,
    lat: float | None = None,
    lon: float | None = None,
    dedupe_timestamps: bool = True,
) -> SeriesV1:
    """Build SeriesV1 from a flat CSV file.

    Accepts ISO-8601 timestamps and common EU day-first formats (e.g. ``23/06/2026 00:00h``).
    Naive timestamps are interpreted in ``timezone`` (IANA, e.g. ``Europe/Madrid``).
    """
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(f"CSV not found: {path}")

    with open(path, newline="", encoding="utf-8-sig") as f:
        return series_from_csv_text(
            f.read(),
            datetime_col=datetime_col,
            columns=columns,
            timezone=timezone,
            source=source,
            site=site,
            lat=lat,
            lon=lon,
            dedupe_timestamps=dedupe_timestamps,
        )


def series_from_excel(
    path: str | Path,
    sheet_name: str | None = None,
    datetime_col: str = "date",
    columns: Sequence[str] | None = None,
    timezone: str = "UTC",
    source: str = "file",
    site: str | None = None,
    lat: float | None = None,
    lon: float | None = None,
    dedupe_timestamps: bool = True,
) -> SeriesV1:
    """Build SeriesV1 from the first worksheet (or ``sheet_name``) of an Excel file."""
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(f"Excel file not found: {path}")

    return series_from_excel_bytes(
        path.read_bytes(),
        sheet_name=sheet_name,
        datetime_col=datetime_col,
        columns=columns,
        timezone=timezone,
        source=source,
        site=site,
        lat=lat,
        lon=lon,
        dedupe_timestamps=dedupe_timestamps,
    )


def wind_series_from_csv(
    path: str | Path,
    datetime_col: str = "date",
    ws_col: str = "ws",
    wd_col: str = "wd",
    timezone: str = "UTC",
    source: str = "file",
    site: str | None = None,
    lat: float | None = None,
    lon: float | None = None,
) -> WindSeriesV1:
    """Build WindSeriesV1 from a flat CSV file."""
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(f"CSV not found: {path}")

    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    if not rows:
        raise ValueError(f"CSV is empty: {path}")

    for required in (datetime_col, ws_col, wd_col):
        if required not in rows[0]:
            raise ValueError(f"Column '{required}' not found. Available: {list(rows[0].keys())}")

    reserved = {datetime_col, ws_col, wd_col}
    pollutant_cols = [c for c in rows[0] if c not in reserved]
    if not pollutant_cols:
        raise ValueError(f"No pollutant columns found (have: {list(rows[0].keys())})")

    timestamps: list[str] = []
    ws_vals: list[float] = []
    wd_vals: list[float] = []
    columns: dict[str, list[float | None]] = {col: [] for col in pollutant_cols}

    for row in rows:
        timestamps.append(parse_series_timestamp(row[datetime_col], source_timezone=timezone))
        try:
            ws_vals.append(float(row[ws_col]))
        except (ValueError, TypeError):
            ws_vals.append(0.0)
        try:
            wd_vals.append(float(row[wd_col]))
        except (ValueError, TypeError):
            wd_vals.append(0.0)
        for col in pollutant_cols:
            raw = (row.get(col) or "").strip()
            try:
                columns[col].append(float(raw))
            except (ValueError, TypeError):
                columns[col].append(None)

    series = [
        SeriesColumnV1(name=normalize_label(col), unit="", values=vals)
        for col, vals in columns.items()
    ]

    return WindSeriesV1(
        timestamps=timestamps,
        series=series,
        ws=ws_vals,
        wd=wd_vals,
        meta=SeriesMetaV1(source=source, site=site, timezone=timezone, lat=lat, lon=lon),
    )
